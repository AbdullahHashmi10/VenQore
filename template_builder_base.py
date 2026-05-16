"""
VenQore — Import Template Builder
==================================
Base code for building any import template with consistent
professional styling. Used as reference by the IDE to generate
new templates matching the VenQore design system.

Usage:
    python template_builder_base.py
    → outputs: <name>_import_template.xlsx

To build a new template:
1. Define COLUMNS list (name, required, guide text, width)
2. Define EXAMPLE_ROWS list
3. Define DATA_VALIDATIONS list
4. Define INSTRUCTIONS list
5. Call build_template()
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ══════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM — All colors, fonts, and spacing in one place.
# Change here and it updates everywhere.
# ══════════════════════════════════════════════════════════════════════════════

COLORS = {
    # Backgrounds
    'title_bg':       '1E293B',   # Dark navy — title banner
    'req_header_bg':  '2563EB',   # Blue — required column header
    'opt_header_bg':  '64748B',   # Grey — optional column header
    'guide_bg':       'FFFBEB',   # Amber — guide row background
    'example_bg_1':   'F0FDF4',   # Light green — example row (odd)
    'example_bg_2':   'ECFDF5',   # Lighter green — example row (even)
    'data_row_1':     'FFFFFF',   # White — data row (even)
    'data_row_2':     'F8FAFC',   # Near white — data row (odd)
    'separator_bg':   '16A34A',   # Green — separator banner
    'section_req':    'EFF6FF',   # Light blue — required field instruction
    'section_opt':    'F8FAFC',   # Near white — optional field instruction
    'section_warn':   'FFFBEB',   # Amber — warning instruction
    'section_ok':     'F0FDF4',   # Light green — tip instruction
    'section_hdr':    '1E293B',   # Dark navy — section header

    # Text
    'white':          'FFFFFF',
    'dark':           '1E293B',
    'guide_text':     '92400E',   # Dark amber
    'example_text':   '166534',   # Dark green
    'req_text':       '1E3A8A',   # Dark blue
    'opt_text':       '334155',   # Slate
    'warn_text':      '92400E',   # Dark amber
    'ok_text':        '166534',   # Dark green

    # Borders
    'border_default': 'CBD5E1',
    'border_guide':   'FCD34D',
    'border_example': 'BBF7D0',
    'border_white':   'FFFFFF',
}

FONTS = {
    'title':        {'name': 'Arial', 'bold': True,  'size': 14},
    'header':       {'name': 'Arial', 'bold': True,  'size': 10},
    'guide':        {'name': 'Arial', 'italic': True,'size': 9},
    'example':      {'name': 'Arial', 'bold': False, 'size': 10},
    'data':         {'name': 'Arial', 'bold': False, 'size': 10},
    'separator':    {'name': 'Arial', 'bold': True,  'size': 9},
    'instruction':  {'name': 'Arial', 'bold': False, 'size': 10},
    'section_hdr':  {'name': 'Arial', 'bold': True,  'size': 10},
}

ROW_HEIGHTS = {
    'title':      32,
    'header':     36,
    'guide':      40,
    'separator':  20,
    'example':    20,
    'data':       18,
    'instruction': None,   # auto-calculated from line count
}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def make_side(color=None):
    return Side(style='thin', color=color or COLORS['border_default'])

def make_border(color=None):
    s = make_side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def make_font(style_key, color=None):
    cfg = FONTS[style_key].copy()
    if color:
        cfg['color'] = color
    return Font(**cfg)

def center_align(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def top_left_align(wrap=True):
    return Alignment(horizontal='left', vertical='top', wrap_text=wrap)

def auto_row_height(text, min_height=20, per_line=15, padding=8):
    """Calculate row height based on number of newlines in text."""
    lines = (text or '').count('\n') + 1
    return max(min_height, lines * per_line + padding)

def instruction_fg(bg_color):
    """Return the correct foreground color for a given background."""
    mapping = {
        COLORS['section_hdr']:  COLORS['white'],
        COLORS['section_req']:  COLORS['req_text'],
        COLORS['section_opt']:  COLORS['opt_text'],
        COLORS['section_warn']: COLORS['warn_text'],
        COLORS['section_ok']:   COLORS['ok_text'],
        'FEF3C7':               COLORS['warn_text'],
        'FFFFFF':               COLORS['opt_text'],
    }
    return mapping.get(bg_color, COLORS['dark'])


# ══════════════════════════════════════════════════════════════════════════════
# CORE BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_sheet1(ws, config):
    """
    Build the main data entry sheet.

    config keys:
        title       (str)  — Banner title text
        columns     (list) — See COLUMN SPEC below
        examples    (list) — See EXAMPLE SPEC below
        data_rows   (int)  — Number of blank data rows (default 300)

    COLUMN SPEC — each column is a dict:
        {
            'label':    'Product Name *',   # Header text
            'required': True,               # Blue header if True, grey if False
            'guide':    'Full name...',     # Guide row text (supports \n)
            'width':    28,                 # Column width
        }

    EXAMPLE SPEC — each example is a list of values in column order:
        ['Pepsi 500ml', 'PEP-001', 120, 80, 'pcs', ...]
    """
    title    = config['title']
    columns  = config['columns']
    examples = config.get('examples', [])
    n_rows   = config.get('data_rows', 300)
    n_cols   = len(columns)

    last_col = get_column_letter(n_cols)

    # ── ROW 1: Title banner ──────────────────────────────────────────────────
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value     = title
    c.font      = make_font('title', COLORS['white'])
    c.fill      = make_fill(COLORS['title_bg'])
    c.alignment = center_align()
    ws.row_dimensions[1].height = ROW_HEIGHTS['title']

    # ── ROW 2: Column headers ────────────────────────────────────────────────
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        c = ws[f'{letter}2']
        c.value     = col['label']
        c.font      = make_font('header', COLORS['white'])
        c.fill      = make_fill(COLORS['req_header_bg'] if col['required']
                                else COLORS['opt_header_bg'])
        c.alignment = center_align()
        c.border    = make_border(COLORS['border_white'])
        ws.column_dimensions[letter].width = col.get('width', 18)
    ws.row_dimensions[2].height = ROW_HEIGHTS['header']

    # ── ROW 3: Guide row ─────────────────────────────────────────────────────
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        c = ws[f'{letter}3']
        c.value     = col.get('guide', '')
        c.font      = make_font('guide', COLORS['guide_text'])
        c.fill      = make_fill(COLORS['guide_bg'])
        c.alignment = center_align()
        c.border    = make_border(COLORS['border_guide'])
    ws.row_dimensions[3].height = ROW_HEIGHTS['guide']

    # ── ROW 4: Separator ─────────────────────────────────────────────────────
    ws.merge_cells(f'A4:{last_col}4')
    c = ws['A4']
    c.value     = '▼  Enter your data below — Row 5 onwards  ▼'
    c.font      = make_font('separator', COLORS['white'])
    c.fill      = make_fill(COLORS['separator_bg'])
    c.alignment = center_align()
    ws.row_dimensions[4].height = ROW_HEIGHTS['separator']

    # ── ROWS 5+: Example rows ─────────────────────────────────────────────────
    for ex_idx, example in enumerate(examples):
        row_num = 5 + ex_idx
        bg = COLORS['example_bg_1'] if ex_idx % 2 == 0 else COLORS['example_bg_2']
        for col_idx, val in enumerate(example, start=1):
            letter = get_column_letter(col_idx)
            c = ws[f'{letter}{row_num}']
            c.value     = val
            c.font      = make_font('example', COLORS['example_text'])
            c.fill      = make_fill(bg)
            c.alignment = left_align()
            c.border    = make_border(COLORS['border_example'])
        ws.row_dimensions[row_num].height = ROW_HEIGHTS['example']

    # ── Blank data rows ───────────────────────────────────────────────────────
    data_start = 5 + len(examples)
    data_end   = data_start + n_rows
    for row_num in range(data_start, data_end):
        bg = COLORS['data_row_1'] if row_num % 2 == 0 else COLORS['data_row_2']
        for col_idx in range(1, n_cols + 1):
            letter = get_column_letter(col_idx)
            c = ws[f'{letter}{row_num}']
            c.fill      = make_fill(bg)
            c.font      = make_font('data', COLORS['dark'])
            c.alignment = left_align()
            c.border    = make_border()
        ws.row_dimensions[row_num].height = ROW_HEIGHTS['data']

    # Freeze header rows
    ws.freeze_panes = 'A5'

    return data_end  # return last row so validations know the range


def add_validation(ws, validation_config, data_end):
    """
    Add data validation to a column range.

    validation_config keys:
        col         (str)  — Column letter e.g. 'B'
        type        (str)  — 'list', 'decimal', 'whole'
        formula     (str)  — e.g. '"Customer,Supplier"' or '0'
        operator    (str)  — for decimal/whole: 'greaterThanOrEqual' etc.
        allow_blank (bool)
        error_title (str)
        error_msg   (str)
        prompt_title(str)
        prompt_msg  (str)
    """
    col      = validation_config['col']
    vtype    = validation_config['type']
    formula  = validation_config['formula']
    operator = validation_config.get('operator')

    kwargs = dict(
        type            = vtype,
        allow_blank     = validation_config.get('allow_blank', True),
        showDropDown    = False,
        showErrorMessage= True,
        errorTitle      = validation_config.get('error_title', 'Invalid Value'),
        error           = validation_config.get('error_msg', 'Please enter a valid value.'),
        showInputMessage= True,
        promptTitle     = validation_config.get('prompt_title', ''),
        prompt          = validation_config.get('prompt_msg', ''),
    )

    if vtype == 'list':
        kwargs['formula1'] = formula
    elif vtype in ('decimal', 'whole'):
        kwargs['operator'] = operator
        kwargs['formula1'] = formula

    dv = DataValidation(**kwargs)
    dv.sqref = f'{col}5:{col}{data_end}'
    ws.add_data_validation(dv)


def build_sheet2(ws, config):
    """
    Build the instructions sheet.

    config keys:
        title        (str)  — Sheet title
        instructions (list) — See INSTRUCTION SPEC below

    INSTRUCTION SPEC — each row is a dict:
        {
            'label':       'Product Name *',
            'explanation': 'Full name...\nSupports newlines.',
            'bg':          COLORS['section_req'],   # or 'section_opt', 'section_warn', 'section_ok', 'section_hdr'
            'bold':        False,
        }
    """
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 72

    # Title
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value     = config['title']
    c.font      = make_font('title', COLORS['white'])
    c.fill      = make_fill(COLORS['title_bg'])
    c.alignment = center_align()
    ws.row_dimensions[1].height = ROW_HEIGHTS['title']

    for row_num, row in enumerate(config['instructions'], start=2):
        bg   = row.get('bg', COLORS['section_opt'])
        bold = row.get('bold', False)
        fg   = instruction_fg(bg)
        expl = row.get('explanation', '')

        a = ws.cell(row=row_num, column=1, value=row.get('label', ''))
        b = ws.cell(row=row_num, column=2, value=expl)

        for c in [a, b]:
            c.font      = Font(name='Arial', bold=bold, size=10, color=fg)
            c.fill      = make_fill(bg)
            c.alignment = top_left_align()
            c.border    = make_border()

        ws.row_dimensions[row_num].height = auto_row_height(expl)

    ws.freeze_panes = 'A2'


def build_sheet3(ws, config):
    """
    Build the valid options reference sheet.

    config keys:
        title   (str)  — Sheet title
        options (list) — See OPTION SPEC below

    OPTION SPEC — each row is a dict:
        {
            'field': 'Base Unit',
            'value': 'pcs',
            'note':  'Individual pieces',
            'bg':    COLORS['section_req'],
            'bold':  False,
        }
    """
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 46

    # Title
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value     = config['title']
    c.font      = make_font('title', COLORS['white'])
    c.fill      = make_fill(COLORS['title_bg'])
    c.alignment = center_align()
    ws.row_dimensions[1].height = ROW_HEIGHTS['title']

    for row_num, row in enumerate(config['options'], start=2):
        bg   = row.get('bg', COLORS['section_opt'])
        bold = row.get('bold', False)
        fg   = instruction_fg(bg)

        for col_idx, text in enumerate(
            [row.get('field',''), row.get('value',''), row.get('note','')],
            start=1
        ):
            c = ws.cell(row=row_num, column=col_idx, value=text)
            c.font      = Font(name='Arial', bold=bold, size=10, color=fg)
            c.fill      = make_fill(bg)
            c.alignment = left_align()
            c.border    = make_border()
        ws.row_dimensions[row_num].height = 22

    ws.freeze_panes = 'A2'


def build_template(filename, sheet1_config, sheet2_config, sheet3_config,
                   validations=None):
    """
    Main entry point. Builds a complete 3-sheet import template.

    Parameters:
        filename        (str)  — Output filename e.g. 'contacts_import_template.xlsx'
        sheet1_config   (dict) — Config for data entry sheet (see build_sheet1)
        sheet2_config   (dict) — Config for instructions sheet (see build_sheet2)
        sheet3_config   (dict) — Config for valid options sheet (see build_sheet3)
        validations     (list) — List of validation configs (see add_validation)
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = '📋 Import Data'
    data_end = build_sheet1(ws1, sheet1_config)

    if validations:
        for v in validations:
            add_validation(ws1, v, data_end)

    ws2 = wb.create_sheet('📖 How To Fill This')
    build_sheet2(ws2, sheet2_config)

    ws3 = wb.create_sheet('✅ Valid Options')
    build_sheet3(ws3, sheet3_config)

    wb.save(filename)
    print(f'✅ Saved: {filename}')


# ══════════════════════════════════════════════════════════════════════════════
# EXAMPLE USAGE — Contacts Import Template
# Delete or replace this section when building a different template.
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':

    sheet1 = {
        'title': 'VenQore — Contacts Import Template',
        'columns': [
            {'label': 'Contact Name *',                 'required': True,  'guide': 'Full name of person or business',              'width': 28},
            {'label': 'Contact Type *',                 'required': True,  'guide': '"Customer" or "Supplier"\nUse the dropdown',   'width': 16},
            {'label': 'Opening Balance',                'required': False, 'guide': 'Numbers only e.g. 5000\nLeave blank if none',  'width': 18},
            {'label': 'Balance Direction *\n(if balance > 0)', 'required': True, 'guide': '"To Receive" → they owe you\n"To Pay" → you owe them', 'width': 22},
            {'label': 'Phone Number',                   'required': False, 'guide': '+923001234567',                                'width': 20},
            {'label': 'Email',                          'required': False, 'guide': 'email@example.com',                           'width': 26},
            {'label': 'Address',                        'required': False, 'guide': 'City or full address',                        'width': 24},
            {'label': 'Credit Limit',                   'required': False, 'guide': 'Max credit allowed\nLeave blank for no limit', 'width': 16},
            {'label': 'Notes',                          'required': False, 'guide': 'Any extra notes',                             'width': 24},
        ],
        'examples': [
            ['Ahmed Traders',    'Customer', 15000, 'To Receive', '+923001234567', 'ahmed@email.com', 'Lahore',     20000, 'Regular buyer'],
            ['Basra Suppliers',  'Supplier', 8500,  'To Pay',     '+923119876543', '',                'Karachi',    '',    '30 day terms'],
            ['Zara Khan',        'Customer', '',    '',           '+923331112233', '',                'Islamabad',  '',    ''],
            ['City Electronics', 'Supplier', '',    '',           '',              'city@elec.com',   'Rawalpindi', '',    ''],
            ['Muhammad Ali Sb',  'Customer', 45000, 'To Receive', '+923456789012', '',                'Faisalabad', 50000, 'VIP customer'],
        ],
        'data_rows': 300,
    }

    sheet2 = {
        'title': 'How to Fill the Contacts Import Template',
        'instructions': [
            {'label': 'COLUMN',              'explanation': 'WHAT TO ENTER',                                                   'bg': COLORS['section_hdr'], 'bold': True},
            {'label': 'Contact Name *',      'explanation': 'Full name. Example: "Ahmed Traders" or "Muhammad Ali Sb"',       'bg': COLORS['section_req'], 'bold': False},
            {'label': 'Contact Type *',      'explanation': '"Customer" = they buy from you.\n"Supplier" = you buy from them.\nUse the dropdown only.', 'bg': COLORS['section_req'], 'bold': False},
            {'label': 'Opening Balance',     'explanation': 'Amount owed before you start. Numbers only — no Rs, no commas.\nLeave blank if none.', 'bg': COLORS['section_opt'], 'bold': False},
            {'label': 'Balance Direction *', 'explanation': '"To Receive" → they owe YOU money.\n"To Pay" → YOU owe them money.\nLeave blank if Opening Balance is 0.', 'bg': COLORS['section_req'], 'bold': False},
            {'label': '',                    'explanation': '',                                                                 'bg': 'FFFFFF',              'bold': False},
            {'label': '⚠️  COMMON MISTAKES', 'explanation': '',                                                                'bg': 'FEF3C7',              'bold': True},
            {'label': 'Wrong Type',          'explanation': 'Only "Customer" or "Supplier" accepted. Not "Client" or "Vendor".','bg': COLORS['section_warn'],'bold': False},
            {'label': 'Rs in balance',       'explanation': 'Enter 5000 not "Rs 5,000". The symbol causes an import error.',   'bg': COLORS['section_warn'],'bold': False},
            {'label': '',                    'explanation': '',                                                                 'bg': 'FFFFFF',              'bold': False},
            {'label': '✅  QUICK RULES',     'explanation': '',                                                                 'bg': COLORS['section_ok'],  'bold': True},
            {'label': '★ = Required',        'explanation': 'Contact Name, Contact Type, and Balance Direction (when balance > 0) are required.', 'bg': COLORS['section_ok'], 'bold': False},
            {'label': 'Example rows',        'explanation': 'Rows 5–9 are examples. Delete them before importing.',           'bg': COLORS['section_ok'],  'bold': False},
        ],
    }

    sheet3 = {
        'title': 'Valid Options — Reference Sheet',
        'options': [
            {'field': 'FIELD',          'value': 'VALID VALUES',  'note': 'WHEN TO USE',                         'bg': COLORS['section_hdr'], 'bold': True},
            {'field': 'Contact Type',   'value': 'Customer',      'note': 'They buy from you',                   'bg': COLORS['section_req'], 'bold': False},
            {'field': '',               'value': 'Supplier',      'note': 'You buy from them',                   'bg': COLORS['section_req'], 'bold': False},
            {'field': 'Balance Direction','value':'To Receive',   'note': 'They owe YOU money',                  'bg': COLORS['section_opt'], 'bold': False},
            {'field': '',               'value': 'To Pay',        'note': 'YOU owe them money',                  'bg': COLORS['section_opt'], 'bold': False},
            {'field': '',               'value': '(leave blank)', 'note': 'Opening Balance is 0 or empty',       'bg': COLORS['section_opt'], 'bold': False},
            {'field': 'Opening Balance','value': 'Any number',    'note': 'e.g. 5000 — numbers only, no Rs',     'bg': COLORS['section_ok'],  'bold': False},
            {'field': '',               'value': '0 or blank',    'note': 'No existing debt',                    'bg': COLORS['section_ok'],  'bold': False},
        ],
    }

    validations = [
        {
            'col': 'B', 'type': 'list', 'formula': '"Customer,Supplier"',
            'allow_blank': False,
            'error_title': 'Invalid Type', 'error_msg': 'Select Customer or Supplier from the dropdown.',
            'prompt_title': 'Contact Type', 'prompt_msg': 'Customer = buys from you.\nSupplier = you buy from them.',
        },
        {
            'col': 'D', 'type': 'list', 'formula': '"To Receive,To Pay"',
            'allow_blank': True,
            'error_title': 'Invalid Direction', 'error_msg': '"To Receive" = they owe you.\n"To Pay" = you owe them.',
            'prompt_title': 'Balance Direction', 'prompt_msg': 'To Receive → they owe you.\nTo Pay → you owe them.\nLeave blank if balance is 0.',
        },
        {
            'col': 'C', 'type': 'decimal', 'operator': 'greaterThanOrEqual', 'formula': '0',
            'allow_blank': True,
            'error_title': 'Invalid Balance', 'error_msg': 'Enter a positive number. No Rs symbol or commas.',
        },
    ]

    build_template(
        filename    = 'contacts_import_template.xlsx',
        sheet1_config = sheet1,
        sheet2_config = sheet2,
        sheet3_config = sheet3,
        validations   = validations,
    )
